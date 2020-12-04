import urllib
import os
import json

import requests
from openpyxl import load_workbook

def get_file_name(url):
    path=urllib.parse.urlparse(url).path
    return os.path.split(path)[-1]

def download_excel_file():
    #奈良県の
    url="http://www.pref.nara.jp/secure/227221/%E5%A5%88%E8%89%AF%E7%9C%8C_02%E6%96%B0%E5%9E%8B%E3%82%B3%E3%83%AD%E3%83%8A%E3%82%A6%E3%82%A4%E3%83%AB%E3%82%B9%E6%84%9F%E6%9F%93%E8%80%85_%E6%82%A3%E8%80%85%E9%9B%86%E8%A8%88%E8%A1%A8.xlsx"
    excel_file_name=get_file_name(url)

    response=requests.get(url)
    if response.status_code!=requests.codes.ok:
        raise Exception("status_code!=200")

    with open(excel_file_name,"wb") as f:
        f.write(response.content)
    return excel_file_name


def main():
    file_name=download_excel_file()

    wb=load_workbook(file_name,data_only=True)
    ws=wb["奈良県_02新型コロナウイルス感染者_患者集計表"]
    
    max_row=ws.max_row
    d=[]
    for i in range(3,max_row):
        dt=ws.cell(i,1).value
        
        if dt is None:
            break
        date=dt.date().isoformat()
        count=ws.cell(i,4).value
        d.append({
            "date":date,
            "count":count
        })

    data={
        "data":d
    }

    with open("nara_data.json","w") as f:
        json.dump(data,f,indent=4)

if __name__ == "__main__":
    main()